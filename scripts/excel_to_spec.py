#!/usr/bin/env python3
"""
excel_to_spec.py — Đọc file test case Excel (mẫu KỊCH BẢN NGHIỆM THU / UAT)
và sinh khung spec Playwright + TypeScript, kèm bảng truy vết test-map.json.

Script tự dò dòng tiêu đề và các cột (tiếng Việt có dấu, không dấu, hoặc tiếng Anh),
xử lý ô gộp (merged cell) và các test case trải trên nhiều dòng.

Chạy `python excel_to_spec.py --help` để xem hướng dẫn.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path

# Console Windows mặc định dùng cp1252, không in được tiếng Việt có dấu và sẽ
# ném UnicodeEncodeError giữa chừng. Ép UTF-8 để script chạy được ở mọi terminal.
for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

try:
    import openpyxl
except ImportError:
    sys.exit(
        "Thiếu thư viện openpyxl.\n"
        "Cài đặt:  pip install openpyxl"
    )


# --------------------------------------------------------------------------
# Nhận diện cột
# --------------------------------------------------------------------------

def normalize(text) -> str:
    """Bỏ dấu tiếng Việt, hạ chữ thường, gộp khoảng trắng — để so khớp tên cột."""
    if text is None:
        return ""
    s = str(text).replace("đ", "d").replace("Đ", "D")
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", s).strip().lower()


# (tên field, danh sách mẫu, điểm ưu tiên) — mẫu càng đặc trưng, điểm càng cao
COLUMN_PATTERNS: list[tuple[str, list[str], int]] = [
    # Mẫu KBKTCN (Kịch bản kiểm thử chức năng) — đặt trước để thắng các mẫu chung
    ("bug_id",       ["id bug", "bug id", "ma loi", "ma bug"], 12),
    ("purpose",      ["muc dich kiem thu", "muc dich"], 12),
    ("priority",     ["thu tu uu tien", "do uu tien", "muc do uu tien", "priority"], 12),
    ("severity",     ["muc do nghiem trong", "severity"], 12),
    ("title",        ["truong hop kiem thu"], 12),
    ("status",       ["ket qua test (m)", "ket qua test (at)", "ket qua test"], 12),

    ("id",           ["ma test case", "ma tc", "test case id", "testcase id", "case id", "ma kich ban"], 10),
    ("id",           ["ma", "id", "stt", "no."], 3),
    ("title",        ["ten test case", "muc tieu", "tieu de", "noi dung kiem tra", "mo ta test case",
                      "ten kich ban", "noi dung kich ban"], 10),
    ("title",        ["mo ta", "noi dung", "chuc nang", "title", "description", "summary", "scenario",
                      "test case", "testcase", "kich ban", "truong hop kiem thu"], 4),
    ("precondition", ["tien dieu kien", "dieu kien tien quyet", "precondition", "pre-condition", "pre condition"], 10),
    ("precondition", ["dieu kien"], 4),
    ("steps",        ["cac buoc thuc hien", "buoc thuc hien", "cac buoc", "test steps", "steps to reproduce"], 10),
    ("steps",        ["thao tac", "steps", "procedure", "thuc hien"], 4),
    ("data",         ["du lieu dau vao", "du lieu test", "test data", "input data"], 10),
    ("data",         ["du lieu", "input", "data"], 4),
    ("expected",     ["ket qua mong doi", "ket qua ky vong", "ket qua du kien", "ket qua mong muon",
                      "expected result", "expected output"], 10),
    ("expected",     ["expected", "ket qua can dat"], 4),
    ("actual",       ["ket qua thuc te", "actual result", "actual"], 10),
    ("status",       ["trang thai", "ket luan", "dat/khong dat", "pass/fail", "status", "result"], 8),
    ("note",         ["ghi chu", "note", "remark", "comment"], 8),
]

FIELDS = ["id", "title", "precondition", "steps", "data", "expected", "actual", "status", "note",
          "bug_id", "purpose", "priority", "severity"]

# Thứ tự ưu tiên trong Excel → tag Playwright, để lọc suite theo mức khi chạy CI.
PRIORITY_TAG = {
    "highest": "@p0", "high": "@p0",
    "medium": "@p1",
    "low": "@p2", "lowest": "@p2", "none": "@p2",
}


def match_column(header: str) -> list[tuple[str, int]]:
    """Trả về các (field, score) khớp với một ô tiêu đề."""
    h = normalize(header)
    if not h:
        return []
    hits = []
    for field_name, patterns, score in COLUMN_PATTERNS:
        for p in patterns:
            if h == p:
                hits.append((field_name, score + 5))
                break
            if p in h:
                hits.append((field_name, score))
                break
    return hits


def map_columns(row: list) -> tuple[dict[str, int], int]:
    """
    Gán cột cho từng field theo kiểu tham lam: cặp (field, cột) nào khớp chắc nhất
    được ưu tiên, và mỗi cột chỉ thuộc về một field.

    Cần thiết vì tên cột hay chồng nhau: 'Kết quả mong đợi' khớp cả `expected` (10)
    lẫn `status` (8, do chứa chữ 'result'), còn 'Mã test case' khớp cả `id` lẫn
    `title`. Không xử lý thì một cột bị hai field cùng nhận và dữ liệu ra sai chỗ.
    """
    matches: list[tuple[int, str, int]] = []
    for col, cell in enumerate(row):
        for field_name, score in match_column(cell):
            matches.append((score, field_name, col))

    matches.sort(key=lambda m: -m[0])

    assigned: dict[str, int] = {}
    used_cols: set[int] = set()
    total = 0
    for score, field_name, col in matches:
        if field_name in assigned or col in used_cols:
            continue
        assigned[field_name] = col
        used_cols.add(col)
        total += score

    return assigned, total


def detect_header_row(rows: list[list], max_scan: int = 20) -> tuple[int, dict[str, int]]:
    """Tìm dòng tiêu đề: dòng khớp được nhiều tên cột đã biết nhất."""
    best_row, best_map, best_score = -1, {}, 0

    for idx, row in enumerate(rows[:max_scan]):
        assigned, score = map_columns(row)

        # Không có cột nội dung nào thì không phải dòng tiêu đề
        if not ({"title", "steps", "expected"} & assigned.keys()):
            continue

        if score > best_score:
            best_row, best_score, best_map = idx, score, assigned

    return best_row, best_map


def header_band_end(rows: list[list], header_row: int, max_band: int = 4) -> int:
    """
    Chỉ số dòng cuối của dải tiêu đề (0-based).

    Mẫu KBKTCN có tiêu đề trải **ba dòng** gộp ô (15–17): dòng 15 là tên cột, 16–17
    là tiêu đề con ('Android (…)', 'Lần 1'…). Sau khi bung ô gộp, dòng 16 và 17 mang
    lại phần lớn giá trị của dòng 15, nên chúng vẫn 'khớp' bộ nhận diện cột — nếu coi
    dòng ngay sau tiêu đề là dữ liệu thì các dòng con này biến thành test case rác.

    Dòng nào còn khớp ≥ 60% điểm của dòng tiêu đề thì vẫn thuộc dải tiêu đề.
    """
    if header_row < 0 or header_row >= len(rows):
        return header_row
    _, base = map_columns(rows[header_row])
    if base <= 0:
        return header_row

    end = header_row
    for idx in range(header_row + 1, min(header_row + 1 + max_band, len(rows))):
        _, score = map_columns(rows[idx])
        if score < base * 0.6:
            break
        end = idx
    return end


PREFIX_LABELS = ("ma truong hop kiem thu", "ma test case", "ma tc", "tien to ma")


def sheet_id_prefix(rows: list[list], header_row: int) -> str:
    """
    Đọc tiền tố mã từ khối metadata phía trên dải tiêu đề.

    Mẫu KBKTCN khai 'Mã trường hợp kiểm thử | QLĐH' ở ô riêng, rồi cột ID dùng công
    thức ghép tiền tố đó với số thứ tự. Đọc được thì spec sinh ra mang đúng mã mà
    tester dùng khi log bug, khỏi phải truyền --id-prefix bằng tay.
    """
    for row in rows[:max(header_row, 0)]:
        for col, cell in enumerate(row):
            if normalize(cell) in PREFIX_LABELS:
                for nxt in row[col + 1:col + 4]:
                    val = "" if nxt is None else str(nxt).strip()
                    if val and not val.startswith("<"):
                        return val + "_"
    return ""


# --------------------------------------------------------------------------
# Đọc dữ liệu
# --------------------------------------------------------------------------

@dataclass
class TestCase:
    tc_id: str
    title: str
    precondition: str = ""
    data: str = ""
    steps: list[str] = field(default_factory=list)
    expected: list[str] = field(default_factory=list)
    excel_row: int = 0
    # Mẫu KBKTCN
    purpose: str = ""       # Mục đích kiểm thử — ô gộp dọc, phủ nhiều dòng TH
    priority: str = ""      # Thứ tự ưu tiên
    severity: str = ""      # Mức độ nghiêm trọng
    bug_id: str = ""        # ID BUG / Mã lỗi đã log cho ca này
    group: str = ""         # SUITE GIAO DIỆN / SUITE CHỨC NĂNG
    suite: str = ""         # Suite 1: <tên>

    @property
    def tag(self) -> str:
        return PRIORITY_TAG.get(normalize(self.priority), "")


def read_sheet_values(ws) -> list[list]:
    """Đọc toàn bộ sheet, đổ giá trị của ô gộp ra mọi ô con để không mất dữ liệu."""
    rows = [[cell.value for cell in row] for row in ws.iter_rows()]

    for rng in ws.merged_cells.ranges:
        top_left = ws.cell(row=rng.min_row, column=rng.min_col).value
        if top_left is None:
            continue
        for r in range(rng.min_row - 1, rng.max_row):
            for c in range(rng.min_col - 1, rng.max_col):
                if r < len(rows) and c < len(rows[r]):
                    rows[r][c] = top_left
    return rows


def cell_text(row: list, col: int | None) -> str:
    if col is None or col >= len(row) or row[col] is None:
        return ""
    return str(row[col]).strip()


STEP_SPLIT = re.compile(r"(?:^|\n)\s*(?:\d+[\.\)]|[-–•*]|Bước\s*\d+\s*[:\.]?)\s*", re.IGNORECASE)


def split_steps(text: str) -> list[str]:
    """Tách khối văn bản thành từng bước: theo xuống dòng, số thứ tự, gạch đầu dòng."""
    if not text:
        return []
    parts = [p.strip() for p in STEP_SPLIT.split(text)]
    parts = [p for p in parts if p]
    if len(parts) <= 1:
        parts = [p.strip() for p in re.split(r"[\n;]+", text) if p.strip()]
    return parts or [text.strip()]


def section_label(row: list, cmap: dict[str, int]) -> str:
    """
    Trả về nhãn nếu đây là dòng phân nhóm (SUITE …), ngược lại trả chuỗi rỗng.

    Trong mẫu KBKTCN, dòng suite là một ô gộp ngang C:G. Sau khi bung ô gộp, cả năm
    cột nội dung mang **cùng một chuỗi** — nếu không nhận ra thì nhãn suite bị đọc
    thành 'kết quả mong muốn' của một test case ma.
    """
    cols = [cmap.get(f) for f in ("purpose", "title", "data", "steps", "expected")]
    vals = [cell_text(row, c) for c in cols if c is not None]
    filled = [v for v in vals if v]
    if len(filled) >= 3 and len(set(filled)) == 1:
        return filled[0]
    return ""


def extract_cases(rows: list[list], header_row: int, cmap: dict[str, int],
                  id_prefix: str, band_end: int | None = None) -> list[TestCase]:
    cases: list[TestCase] = []
    counter = 0
    group = suite = ""
    start = (band_end if band_end is not None else header_row) + 1

    # rows[i] là dòng Excel i+1 (0-based → 1-based), nên offset phải là index + 1.
    for offset, row in enumerate(rows[start:], start=start + 1):
        if all(c is None or str(c).strip() == "" for c in row):
            continue

        label = section_label(row, cmap)
        if label:
            # Nhóm lớn viết hoa toàn bộ (SUITE GIAO DIỆN); còn lại là suite con.
            if label == label.upper():
                group, suite = label, ""
            else:
                suite = label
            continue

        raw_id = cell_text(row, cmap.get("id"))
        title = cell_text(row, cmap.get("title"))
        steps = cell_text(row, cmap.get("steps"))
        expected = cell_text(row, cmap.get("expected"))
        precondition = cell_text(row, cmap.get("precondition"))
        data = cell_text(row, cmap.get("data"))
        purpose = cell_text(row, cmap.get("purpose"))
        priority = cell_text(row, cmap.get("priority"))
        severity = cell_text(row, cmap.get("severity"))
        bug_id = cell_text(row, cmap.get("bug_id"))

        # Dòng tiêu đề nhóm (chỉ có chữ ở cột title, không có bước/kết quả) — bỏ qua
        if title and not steps and not expected and not raw_id:
            continue

        starts_new = bool(raw_id) if "id" in cmap else bool(title)
        # Ô gộp làm ID lặp lại ở các dòng con → dòng nào trùng ID dòng trước là dòng tiếp nối.
        # NHƯNG nếu dòng đó có tiêu đề riêng khác dòng trước thì nó là một ca độc lập đang
        # bị TRÙNG mã (công thức sinh ID trong Excel sai). Gộp lại sẽ nuốt mất hẳn một ca —
        # tách ra rồi để audit_ids() báo trùng, đừng im lặng làm mất dữ liệu.
        if starts_new and cases and raw_id and raw_id == cases[-1].tc_id:
            starts_new = bool(title) and title != cases[-1].title

        if starts_new or not cases:
            counter += 1
            tc_id = raw_id or f"{id_prefix}{counter:02d}"
            if re.fullmatch(r"\d+", tc_id):          # cột STT chỉ có số → gắn tiền tố cho dễ đọc
                tc_id = f"{id_prefix}{int(tc_id):02d}"
            cases.append(TestCase(
                tc_id=tc_id,
                title=title or purpose or f"Test case {tc_id}",
                precondition=precondition,
                data=data,
                excel_row=offset,
                purpose=purpose,
                priority=priority,
                severity=severity,
                bug_id=bug_id,
                group=group,
                suite=suite,
            ))

        cur = cases[-1]
        cur.steps.extend(split_steps(steps))
        cur.expected.extend(split_steps(expected))
        for attr, val in (("precondition", precondition), ("data", data),
                          ("purpose", purpose), ("priority", priority),
                          ("severity", severity), ("bug_id", bug_id)):
            if val and not getattr(cur, attr):
                setattr(cur, attr, val)

    return cases


def audit_ids(cases: list[TestCase]) -> list[str]:
    """
    Soát mã test case: trùng mã thì truy vết gãy (một bug không biết thuộc ca nào),
    thủng số thì thường là công thức sinh ID trong Excel bị trôi tham chiếu.
    """
    warnings: list[str] = []
    seen: dict[str, list[int]] = {}
    for c in cases:
        seen.setdefault(c.tc_id, []).append(c.excel_row)
    for tc_id, at in seen.items():
        if len(at) > 1:
            warnings.append(f"Mã '{tc_id}' bị TRÙNG ở dòng Excel {', '.join(map(str, at))}")

    nums = sorted(int(m.group(1)) for c in cases
                  if (m := re.search(r"(\d+)\s*$", c.tc_id)))
    if len(nums) > 2:
        missing = [n for n in range(nums[0], nums[-1] + 1) if n not in nums]
        if missing:
            head = ", ".join(map(str, missing[:10]))
            more = f" (+{len(missing) - 10} nữa)" if len(missing) > 10 else ""
            warnings.append(f"Dãy mã thủng ở số: {head}{more}")
    return warnings


# --------------------------------------------------------------------------
# Sinh code
# --------------------------------------------------------------------------

def slugify(text: str) -> str:
    s = normalize(text)
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return s or "test-cases"


def ts_string(text: str) -> str:
    """Escape chuỗi để nhúng vào code TypeScript (dùng nháy đơn)."""
    return (str(text).replace("\\", "\\\\").replace("'", "\\'")
            .replace("\r", " ").replace("\n", " ").strip())


def block_comment(text: str) -> str:
    return str(text).replace("*/", "*\\/")


def generate_spec(sheet_name: str, cases: list[TestCase], source_file: str) -> str:
    lines = [
        "/**",
        f" * Sinh tự động từ: {block_comment(source_file)} — sheet \"{block_comment(sheet_name)}\"",
        f" * Số test case: {len(cases)}",
        " *",
        " * ⚠ ĐÂY LÀ KHUNG, CHƯA CHẠY ĐƯỢC.",
        " * Cần thay mỗi TODO bằng thao tác và assertion thật.",
        " * Lấy locator có thật bằng:  node scripts/explore.mjs --url <url>",
        " *",
        " * Assertion mặc định để FAIL có chủ ý — một khung test luôn xanh nguy hiểm hơn",
        " * là không có test, vì nó tạo cảm giác đã kiểm tra trong khi chưa kiểm tra gì.",
        " */",
        "import { test, expect } from '@playwright/test';",
        "",
        f"test.describe('{ts_string(sheet_name)}', () => {{",
    ]

    # Gom theo suite để report hiện đúng cấu trúc trong file Excel. Không có suite
    # (mẫu Excel phẳng) thì chỉ một nhóm rỗng, cấu trúc giữ nguyên như trước.
    groups: dict[str, list[TestCase]] = {}
    for case in cases:
        key = " › ".join(x for x in (case.group, case.suite) if x)
        groups.setdefault(key, []).append(case)

    for suite_name, suite_cases in groups.items():
        pad = "  " if suite_name else ""
        if suite_name:
            lines.append("")
            lines.append(f"  test.describe('{ts_string(suite_name)}', () => {{")

        for case in suite_cases:
            lines.append("")
            meta = [m for m in (
                f"Mục đích: {block_comment(case.purpose)}" if case.purpose else "",
                f"Tiền điều kiện: {block_comment(case.precondition)}" if case.precondition else "",
                f"Dữ liệu: {block_comment(case.data)}" if case.data else "",
                f"Ưu tiên: {case.priority}" if case.priority else "",
                f"Mức độ: {case.severity}" if case.severity else "",
                f"Bug đã log: {block_comment(case.bug_id)}" if case.bug_id else "",
            ) if m]
            if meta:
                lines.append(f"{pad}  /**")
                for m in meta:
                    lines.append(f"{pad}   * {m}")
                lines.append(f"{pad}   * Dòng trong Excel: {case.excel_row}")
                lines.append(f"{pad}   */")

            title = f"{case.tc_id}: {case.title}" if not case.title.startswith(case.tc_id) else case.title
            opts = f", {{ tag: '{case.tag}' }}" if case.tag else ""
            lines.append(f"{pad}  test('{ts_string(title)}'{opts}, async ({{ page }}) => {{")

            if case.precondition:
                lines.append(f"{pad}    // Tiền điều kiện: {block_comment(case.precondition)}")
                lines.append(f"{pad}    // TODO: chuẩn bị tiền điều kiện (đăng nhập sẵn, tạo dữ liệu qua API...)")
                lines.append("")

            if not case.steps:
                lines.append(f"{pad}    // TODO: file Excel không mô tả bước thực hiện cho ca này")

            for i, step in enumerate(case.steps, start=1):
                lines.append(f"{pad}    await test.step('{ts_string(f'{i}. {step}')}', async () => {{")
                lines.append(f"{pad}      // TODO: {block_comment(step)}")
                lines.append(f"{pad}    }});")
                lines.append("")

            expected_text = " | ".join(case.expected) if case.expected else "(Excel không ghi kết quả mong đợi)"
            lines.append(f"{pad}    await test.step('Kết quả mong đợi: {ts_string(expected_text)}', async () => {{")
            for exp in case.expected:
                lines.append(f"{pad}      // TODO: assertion cho — {block_comment(exp)}")
            lines.append(f"{pad}      // Ví dụ: await expect(page.getByText('...')).toBeVisible();")
            lines.append(f"{pad}      expect(true, 'Chưa hiện thực assertion cho ca này').toBe(false);")
            lines.append(f"{pad}    }});")
            lines.append(f"{pad}  }});")

        if suite_name:
            lines.append("  });")

    lines.append("});")
    lines.append("")
    return "\n".join(lines)


# --------------------------------------------------------------------------
# CLI
# --------------------------------------------------------------------------

def main() -> int:
    p = argparse.ArgumentParser(
        description="Chuyển file test case Excel thành khung spec Playwright + TypeScript.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
VÍ DỤ
  python excel_to_spec.py --file "KỊCH BẢN NGHIỆM THU.xlsx" --dry-run
  python excel_to_spec.py --file testcase.xlsx --out ./e2e/tests/generated
  python excel_to_spec.py --file testcase.xlsx --sheet "Đăng nhập" --header-row 5
  python excel_to_spec.py --file testcase.xlsx --col-id "STT" --col-expected "Kết quả mong đợi"

QUY TRÌNH ĐẦY ĐỦ
  1. --dry-run           kiểm tra script dò đúng cột chưa
  2. --out ...           sinh khung spec
  3. explore.mjs         lấy locator thật từ app
  4. điền TODO           thay locator + assertion
  5. npx playwright test chạy và sửa
""")
    p.add_argument("--file", required=True, help="Đường dẫn file .xlsx")
    p.add_argument("--out", default="./tests/generated", help="Thư mục xuất spec (mặc định: ./tests/generated)")
    p.add_argument("--sheet", action="append", help="Chỉ xử lý sheet này (lặp lại được). Mặc định: tất cả")
    p.add_argument("--list-sheets", action="store_true", help="Chỉ liệt kê tên các sheet rồi thoát")
    p.add_argument("--dry-run", action="store_true", help="Chỉ in phân tích, không ghi file")
    p.add_argument("--force", action="store_true", help="Ghi đè file spec đã tồn tại")
    p.add_argument("--header-row", type=int, help="Số dòng tiêu đề (1-based), nếu tự dò sai")
    p.add_argument("--id-prefix", default="TC-", help="Tiền tố mã TC khi Excel không có cột mã (mặc định: TC-)")
    for f in FIELDS:
        p.add_argument(f"--col-{f}", help=f"Chỉ định tên cột cho '{f}'")
    args = p.parse_args()

    src = Path(args.file)
    if not src.exists():
        print(f"✗ Không tìm thấy file: {src}", file=sys.stderr)
        return 1

    try:
        wb = openpyxl.load_workbook(src, data_only=True)
    except Exception as exc:
        print(f"✗ Không đọc được file Excel: {exc}", file=sys.stderr)
        return 1

    if args.list_sheets:
        print(f"Các sheet trong {src.name}:")
        for name in wb.sheetnames:
            print(f"  · {name}")
        return 0

    sheets = args.sheet or wb.sheetnames
    missing = [s for s in sheets if s not in wb.sheetnames]
    if missing:
        print(f"✗ Không có sheet: {', '.join(missing)}", file=sys.stderr)
        print(f"  Sheet hiện có: {', '.join(wb.sheetnames)}", file=sys.stderr)
        return 1

    out_dir = Path(args.out)
    overrides = {f: getattr(args, f"col_{f}") for f in FIELDS if getattr(args, f"col_{f}")}
    trace: list[dict] = []
    total_cases = 0

    print(f"\n{'═' * 62}\n File: {src.name}\n Sheet xử lý: {len(sheets)}\n{'═' * 62}")

    for sheet_name in sheets:
        ws = wb[sheet_name]
        rows = read_sheet_values(ws)
        if not rows:
            print(f"\n▸ {sheet_name}: sheet rỗng, bỏ qua")
            continue

        if args.header_row:
            header_row = args.header_row - 1
            # Người dùng đã chỉ định dòng tiêu đề nên không áp điều kiện lọc như khi tự dò.
            cmap = map_columns(rows[header_row])[0] if 0 <= header_row < len(rows) else {}
        else:
            header_row, cmap = detect_header_row(rows)

        # Áp chỉ định tay của người dùng
        if overrides and 0 <= header_row < len(rows):
            header_cells = rows[header_row]
            for fname, wanted in overrides.items():
                target = normalize(wanted)
                for col, cell in enumerate(header_cells):
                    if normalize(cell) == target or (target and target in normalize(cell)):
                        # Gỡ field khác đang chiếm cột này, tránh hai field đọc cùng một cột
                        for other in [f for f, c in cmap.items() if c == col and f != fname]:
                            del cmap[other]
                        cmap[fname] = col
                        break

        if header_row < 0 or not cmap:
            print(f"\n▸ {sheet_name}: ✗ không nhận ra dòng tiêu đề.")
            print("   Dùng --header-row <số dòng> và --col-* để chỉ định tay.")
            print("   10 dòng đầu của sheet:")
            for i, row in enumerate(rows[:10], start=1):
                preview = " | ".join(str(c)[:22] if c is not None else "" for c in row[:8])
                print(f"     {i:>3}: {preview}")
            continue

        # Sheet tổng hợp/dashboard không có cột kết quả mong đợi — nó là báo cáo,
        # không phải danh sách test case. Sinh spec từ nó chỉ ra các ca rác kiểu
        # 'Total — Total'. Bỏ qua, trừ khi người dùng chỉ định dòng tiêu đề bằng tay.
        if "expected" not in cmap and not args.header_row:
            print(f"\n▸ {sheet_name}: bỏ qua — không có cột kết quả mong đợi "
                  f"(nhiều khả năng là sheet tổng hợp, không phải danh sách test case).")
            continue

        band_end = header_band_end(rows, header_row)
        prefix = sheet_id_prefix(rows, header_row) or args.id_prefix
        cases = extract_cases(rows, header_row, cmap, prefix, band_end=band_end)

        header_cells = rows[header_row]
        mapping_desc = ", ".join(
            f"{f}='{str(header_cells[c])[:24]}'" for f, c in sorted(cmap.items(), key=lambda x: x[1])
            if c < len(header_cells) and header_cells[c] is not None
        )
        print(f"\n▸ {sheet_name}")
        print(f"   Dòng tiêu đề: {header_row + 1}")
        print(f"   Cột nhận diện: {mapping_desc or '(không có)'}")
        print(f"   Test case đọc được: {len(cases)}")

        for c in cases[:5]:
            print(f"     · {c.tc_id} — {c.title[:58]} ({len(c.steps)} bước, {len(c.expected)} kết quả)")
        if len(cases) > 5:
            print(f"     ... còn {len(cases) - 5} ca nữa")

        suites = [s for s in dict.fromkeys(f"{c.group} › {c.suite}".strip(" ›") for c in cases) if s]
        if suites:
            print(f"   Suite nhận diện: {len(suites)} — {'; '.join(suites[:3])}"
                  f"{' …' if len(suites) > 3 else ''}")

        missing_expected = [c.tc_id for c in cases if not c.expected]
        if missing_expected:
            print(f"   ⚠ {len(missing_expected)} ca không có kết quả mong đợi: {', '.join(missing_expected[:6])}")
            print("     Ca không có kết quả mong đợi thì không kiểm chứng được — nên bổ sung vào Excel trước.")

        for w in audit_ids(cases):
            print(f"   ⚠ {w}")
            print("     Mã trùng/thủng thường do công thức sinh ID trong Excel bị trôi tham chiếu"
                  " (COUNTBLANK không bám đúng dòng hiện tại). Xem assets/testcase-template/KBKTCN.xlsx.")

        if not cases:
            continue
        total_cases += len(cases)

        spec_name = f"{slugify(sheet_name)}.spec.ts"
        for c in cases:
            trace.append({
                "tc_id": c.tc_id, "sheet": sheet_name, "excel_row": c.excel_row,
                "title": c.title, "steps": len(c.steps),
                "spec_file": spec_name, "status": "generated",
            })

        if args.dry_run:
            continue

        out_dir.mkdir(parents=True, exist_ok=True)
        dest = out_dir / spec_name
        if dest.exists() and not args.force:
            print(f"   ⚠ Đã tồn tại, bỏ qua (dùng --force để ghi đè): {dest}")
            continue
        dest.write_text(generate_spec(sheet_name, cases, src.name), encoding="utf-8")
        print(f"   ✓ Đã ghi: {dest}")

    if not args.dry_run and trace:
        out_dir.mkdir(parents=True, exist_ok=True)
        map_file = out_dir / "test-map.json"
        map_file.write_text(json.dumps(trace, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"\n✓ Bảng truy vết: {map_file}  ({len(trace)} test case)")

    print(f"\n{'─' * 62}")
    if args.dry_run:
        print(f" DRY RUN — đọc được {total_cases} test case, chưa ghi file nào.")
        print(" Nếu bảng cột ở trên đúng, chạy lại và bỏ --dry-run.")
    else:
        print(f" Xong. {total_cases} test case → {out_dir}")
        print(" Bước tiếp theo: node scripts/explore.mjs --url <url>  để lấy locator thật,")
        print(" rồi thay các TODO trong file spec.")
    print(f"{'─' * 62}\n")
    return 0


if __name__ == "__main__":
    sys.exit(main())
